"""Generate the TRA audit findings register as a multi-sheet xlsx.

Sheets:
  1. Summary — counts by severity, track, category
  2. Findings — full register, one row per finding, filterable
  3. Spec Conformance — Track A subset
  4. Code Quality — Track B subset
  5. Doc Consistency — Track C subset
  6. Test Suite — Track D subset
  7. Remediation Backlog — prioritized action items
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/home/z/my-project/scripts")
from tra_findings import FINDINGS, by_category, by_track, stats  # noqa: E402

# ---------------------------------------------------------------------------
# Design tokens (Business Cool palette per xlsx skill design system)
# ---------------------------------------------------------------------------
COLOR_PRIMARY = "243447"
COLOR_BODY = "182030"
COLOR_SECONDARY = "5B6B7D"
COLOR_ACCENT = "4C6EF5"
COLOR_SURFACE = "F8FAFC"
COLOR_ALT_ROW = "F1F5F9"
COLOR_BLOCKING = "C92A2A"
COLOR_WARNING = "F08C00"
COLOR_INFO = "1C7ED6"
COLOR_HEADER_BG = "243447"
COLOR_HEADER_TEXT = "FFFFFF"

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=COLOR_PRIMARY)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_TEXT)
FONT_BODY = Font(name="Calibri", size=10, color=COLOR_BODY)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=COLOR_BODY)
FONT_SEVERITY_BLOCKING = Font(name="Calibri", size=10, bold=True, color=COLOR_BLOCKING)
FONT_SEVERITY_WARNING = Font(name="Calibri", size=10, bold=True, color=COLOR_WARNING)
FONT_SEVERITY_INFO = Font(name="Calibri", size=10, bold=True, color=COLOR_INFO)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_SURFACE = PatternFill("solid", fgColor=COLOR_SURFACE)
FILL_ALT = PatternFill("solid", fgColor=COLOR_ALT_ROW)
FILL_BLOCKING = PatternFill("solid", fgColor="FFF5F5")
FILL_WARNING = PatternFill("solid", fgColor="FFFBEB")
FILL_INFO = PatternFill("solid", fgColor="EFF6FF")

ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_BODY = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def severity_font(sev: str) -> Font:
    if sev == "BLOCKING":
        return FONT_SEVERITY_BLOCKING
    if sev == "WARNING":
        return FONT_SEVERITY_WARNING
    return FONT_SEVERITY_INFO


def severity_fill(sev: str) -> PatternFill:
    if sev == "BLOCKING":
        return FILL_BLOCKING
    if sev == "WARNING":
        return FILL_WARNING
    return FILL_INFO


def write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def set_column_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_zebra_and_borders(ws, start_row: int, end_row: int, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                if (r - start_row) % 2 == 1:
                    cell.fill = FILL_ALT
                cell.font = FONT_BODY
                cell.alignment = ALIGN_BODY


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------
def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "TRA Prototype Audit — Findings Summary"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:E1")

    ws["A3"] = "Audit date:"
    ws["A3"].font = FONT_BODY_BOLD
    ws["B3"] = "2026-07-13"
    ws["B3"].font = FONT_BODY

    ws["A4"] = "Auditor:"
    ws["A4"].font = FONT_BODY_BOLD
    ws["B4"] = "Super Z (multi-agent, 4-track parallel audit)"
    ws["B4"].font = FONT_BODY

    ws["A5"] = "Repo:"
    ws["A5"].font = FONT_BODY_BOLD
    ws["B5"] = "github.com/nordeim/Translation-Runtime-Architecture @ HEAD"
    ws["B5"].font = FONT_BODY

    ws["A6"] = "Severity lexicon:"
    ws["A6"].font = FONT_BODY_BOLD
    ws["B6"] = "TRA-SPECIFICATION.md §7 (BLOCKING / WARNING / INFO)"
    ws["B6"].font = FONT_BODY

    # Totals table
    row = 8
    ws.cell(row=row, column=1, value="Severity").font = FONT_HEADER
    ws.cell(row=row, column=1).fill = FILL_HEADER
    ws.cell(row=row, column=1).alignment = ALIGN_HEADER
    ws.cell(row=row, column=2, value="Count").font = FONT_HEADER
    ws.cell(row=row, column=2).fill = FILL_HEADER
    ws.cell(row=row, column=2).alignment = ALIGN_HEADER
    ws.row_dimensions[row].height = 24

    s = stats()
    for i, sev in enumerate(["BLOCKING", "WARNING", "INFO", "TOTAL"], start=row + 1):
        ws.cell(row=i, column=1, value=sev).font = severity_font(sev) if sev != "TOTAL" else FONT_BODY_BOLD
        ws.cell(row=i, column=2, value=s[sev]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2).border = THIN_BORDER
        ws.cell(row=i, column=1).alignment = ALIGN_BODY
        ws.cell(row=i, column=2).alignment = ALIGN_BODY

    # By track
    row = 15
    ws.cell(row=row, column=1, value="By Track").font = FONT_TITLE
    ws.row_dimensions[row].height = 26
    row += 1
    write_header_row(ws, row, ["Track", "Name", "BLOCKING", "WARNING", "INFO", "TOTAL"])
    row += 1
    track_names = {"A": "Spec Conformance", "B": "Code Quality & Security", "C": "Doc Consistency", "D": "Test Suite"}
    bt = by_track()
    for track_id in ["A", "B", "C", "D"]:
        c = bt[track_id]
        ws.cell(row=row, column=1, value=track_id).font = FONT_BODY_BOLD
        ws.cell(row=row, column=2, value=track_names[track_id]).font = FONT_BODY
        ws.cell(row=row, column=3, value=c["BLOCKING"]).font = FONT_SEVERITY_BLOCKING
        ws.cell(row=row, column=4, value=c["WARNING"]).font = FONT_SEVERITY_WARNING
        ws.cell(row=row, column=5, value=c["INFO"]).font = FONT_SEVERITY_INFO
        ws.cell(row=row, column=6, value=c["TOTAL"]).font = FONT_BODY_BOLD
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = ALIGN_BODY
        row += 1

    # By category
    row += 1
    ws.cell(row=row, column=1, value="By Category").font = FONT_TITLE
    ws.row_dimensions[row].height = 26
    row += 1
    write_header_row(ws, row, ["Category", "BLOCKING", "WARNING", "INFO", "TOTAL"])
    row += 1
    bc = by_category()
    for cat, c in bc.items():
        ws.cell(row=row, column=1, value=cat).font = FONT_BODY
        ws.cell(row=row, column=2, value=c["BLOCKING"]).font = FONT_SEVERITY_BLOCKING
        ws.cell(row=row, column=3, value=c["WARNING"]).font = FONT_SEVERITY_WARNING
        ws.cell(row=row, column=4, value=c["INFO"]).font = FONT_SEVERITY_INFO
        ws.cell(row=row, column=5, value=c["TOTAL"]).font = FONT_BODY_BOLD
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = ALIGN_BODY
        row += 1

    set_column_widths(ws, [22, 32, 14, 14, 14, 14])

    # Freeze panes
    ws.freeze_panes = "A8"


# ---------------------------------------------------------------------------
# Sheet 2: Findings (full register)
# ---------------------------------------------------------------------------
def build_findings_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Findings")

    headers = [
        "ID",
        "Severity",
        "Track",
        "Category",
        "Title",
        "Evidence (file:line)",
        "Detail",
        "Suggested Fix",
    ]
    write_header_row(ws, 1, headers)

    for i, f in enumerate(FINDINGS, start=2):
        ws.cell(row=i, column=1, value=f["id"]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2, value=f["severity"]).font = severity_font(f["severity"])
        ws.cell(row=i, column=3, value=f["track"]).font = FONT_BODY
        ws.cell(row=i, column=4, value=f["category"]).font = FONT_BODY
        ws.cell(row=i, column=5, value=f["title"]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=6, value=f["evidence"]).font = FONT_BODY
        ws.cell(row=i, column=7, value=f["detail"]).font = FONT_BODY
        ws.cell(row=i, column=8, value=f["suggested_fix"]).font = FONT_BODY

        # Severity-colored fill on the severity cell
        ws.cell(row=i, column=2).fill = severity_fill(f["severity"])

        for col in range(1, 9):
            ws.cell(row=i, column=col).border = THIN_BORDER
            if col != 2:
                ws.cell(row=i, column=col).alignment = ALIGN_BODY
            else:
                ws.cell(row=i, column=col).alignment = ALIGN_CENTER
        # Variable row height based on detail length
        ws.row_dimensions[i].height = max(60, min(220, len(f["detail"]) // 4))

    set_column_widths(ws, [10, 11, 7, 28, 38, 30, 70, 55])
    ws.freeze_panes = "A2"

    # Enable autofilter
    ws.auto_filter.ref = f"A1:H{len(FINDINGS) + 1}"


# ---------------------------------------------------------------------------
# Sheets 3-6: Per-track subsets
# ---------------------------------------------------------------------------
def build_track_sheet(wb: Workbook, track_id: str, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    track_findings = [f for f in FINDINGS if f["track"] == track_id]
    headers = ["ID", "Severity", "Category", "Title", "Evidence", "Detail", "Suggested Fix"]
    write_header_row(ws, 1, headers)
    for i, f in enumerate(track_findings, start=2):
        ws.cell(row=i, column=1, value=f["id"]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2, value=f["severity"]).font = severity_font(f["severity"])
        ws.cell(row=i, column=3, value=f["category"]).font = FONT_BODY
        ws.cell(row=i, column=4, value=f["title"]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=5, value=f["evidence"]).font = FONT_BODY
        ws.cell(row=i, column=6, value=f["detail"]).font = FONT_BODY
        ws.cell(row=i, column=7, value=f["suggested_fix"]).font = FONT_BODY
        ws.cell(row=i, column=2).fill = severity_fill(f["severity"])
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = ALIGN_BODY if col != 2 else ALIGN_CENTER
        ws.row_dimensions[i].height = max(60, min(220, len(f["detail"]) // 4))
    set_column_widths(ws, [10, 11, 28, 38, 30, 70, 55])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(track_findings) + 1}"


# ---------------------------------------------------------------------------
# Sheet 7: Remediation Backlog
# ---------------------------------------------------------------------------
def build_backlog_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Remediation Backlog")

    # Priority: BLOCKING first, then WARNING, then INFO; within each, by track
    priority_order = {"BLOCKING": 1, "WARNING": 2, "INFO": 3}
    sorted_findings = sorted(FINDINGS, key=lambda f: (priority_order[f["severity"]], f["track"], f["id"]))

    headers = ["Priority", "ID", "Severity", "Category", "Title", "Suggested Fix", "Estimated Effort"]
    write_header_row(ws, 1, headers)

    effort_estimates = {
        "TRA-001": "Large (1-2 days)",
        "TRA-002": "Medium (4-6h)",
        "TRA-003": "Small (1-2h)",
        "TRA-004": "Medium (4-6h)",
        "TRA-005": "Small (2-3h)",
        "TRA-006": "Large (1-2 days)",
        "TRA-007": "Small (1-2h)",
        "TRA-008": "Small (2-3h)",
        "TRA-009": "Small (1h, doc only)",
        "TRA-010": "Medium (4-6h)",
        "TRA-011": "Small (1-2h)",
        "TRA-012": "Small (2-3h)",
        "TRA-013": "Medium (4-6h)",
        "TRA-014": "Small (2-3h)",
        "TRA-015": "Small (1h)",
        "TRA-016": "Small (30min)",
        "TRA-017": "Small (1-2h)",
        "TRA-018": "Medium (4-6h)",
        "TRA-019": "Small (30min)",
        "TRA-020": "Small (1h, doc only)",
        "TRA-021": "Small (30min, doc only)",
        "TRA-022": "Small (15min, doc only)",
        "TRA-023": "Small (15min, doc only)",
        "TRA-024": "Small (1h, doc only)",
        "TRA-025": "Small (30min)",
        "TRA-026": "Small (30min)",
        "TRA-027": "Small (30min)",
        "TRA-028": "Small (1-2h)",
        "TRA-029": "Small (1h)",
        "TRA-030": "Small (1h)",
        "TRA-031": "Medium (4-6h)",
        "TRA-032": "Small (2-3h)",
        "TRA-033": "Small (1h)",
        "TRA-034": "Small (1-2h)",
        "TRA-035": "N/A (informational)",
    }

    for i, f in enumerate(sorted_findings, start=2):
        priority = f"P{i - 1}"
        ws.cell(row=i, column=1, value=priority).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2, value=f["id"]).font = FONT_BODY_BOLD
        ws.cell(row=i, column=3, value=f["severity"]).font = severity_font(f["severity"])
        ws.cell(row=i, column=4, value=f["category"]).font = FONT_BODY
        ws.cell(row=i, column=5, value=f["title"]).font = FONT_BODY
        ws.cell(row=i, column=6, value=f["suggested_fix"]).font = FONT_BODY
        ws.cell(row=i, column=7, value=effort_estimates.get(f["id"], "TBD")).font = FONT_BODY
        ws.cell(row=i, column=3).fill = severity_fill(f["severity"])
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = ALIGN_BODY if col not in (1, 3) else ALIGN_CENTER
        ws.row_dimensions[i].height = max(40, min(160, len(f["suggested_fix"]) // 4))

    set_column_widths(ws, [9, 10, 11, 28, 38, 60, 18])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(sorted_findings) + 1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    wb = Workbook()
    wb.properties.creator = "Z.ai"
    wb.properties.title = "TRA Prototype Audit — Findings Register"
    wb.properties.subject = "Code review & audit of Translation-Runtime-Architecture"

    build_summary_sheet(wb)
    build_findings_sheet(wb)
    build_track_sheet(wb, "A", "Track A - Spec Conformance")
    build_track_sheet(wb, "B", "Track B - Code Quality")
    build_track_sheet(wb, "C", "Track C - Doc Consistency")
    build_track_sheet(wb, "D", "Track D - Test Suite")
    build_backlog_sheet(wb)

    out_path = Path("/home/z/my-project/download/TRA_audit_findings_register.xlsx")
    wb.save(out_path)
    print(f"XLSX saved to {out_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
