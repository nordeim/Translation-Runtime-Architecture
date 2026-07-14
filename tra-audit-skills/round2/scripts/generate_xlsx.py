"""Generate the multi-sheet XLSX findings register for Round 2 audit.

Sheets:
- Summary: counts by severity, track, category, round-1 status
- Findings: full 41-row register with all columns
- Track A-E: per-track subsets
- Round1 Status: carry-over vs new
- Remediation Backlog: priority-sorted with effort estimates
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REGISTER_PATH = Path("/home/z/my-project/audit-ctx/master_findings_register.json")
OUT_PATH = Path("/home/z/my-project/download/TRA_audit_findings_register_r2.xlsx")

# Severity color coding
SEVERITY_FILL = {
    "BLOCKING": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "WARNING": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "INFO": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="263238"),
    top=Side(style="thin", color="263238"),
    left=Side(style="thin", color="263238"),
    right=Side(style="thin", color="263238"),
)

# Effort estimate by severity (person-hours)
EFFORT_HOURS = {"BLOCKING": 4, "WARNING": 2, "INFO": 0.5}

# Columns for the Findings sheet
FINDINGS_COLUMNS = [
    ("id", "ID", 10),
    ("severity", "Severity", 10),
    ("category", "Category", 35),
    ("track", "Track", 8),
    ("title", "Title", 50),
    ("evidence", "Evidence", 60),
    ("detail", "Detail", 80),
    ("suggested_fix", "Suggested Fix", 60),
    ("round1_status", "Round 1 Status", 14),
    ("source_findings", "Source Findings", 30),
]


def style_header(ws, columns: list[tuple[str, str, int]]) -> None:
    for idx, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 30


def write_finding_row(ws, row_idx: int, finding: dict, columns: list[tuple[str, str, int]]) -> None:
    for col_idx, (key, _, _) in enumerate(columns, 1):
        value = finding.get(key, "")
        if isinstance(value, list):
            value = "; ".join(value)
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if key == "severity":
            cell.fill = SEVERITY_FILL.get(value, PatternFill())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    findings = json.loads(REGISTER_PATH.read_text(encoding="utf-8"))
    wb = Workbook()

    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "TRA Prototype Audit Round 2 — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Repo: nordeim/Translation-Runtime-Architecture @ HEAD 4b8827c"
    ws["A3"] = f"Audit date: 2026-07-14"
    ws["A4"] = f"Total findings: {len(findings)}"

    # By severity
    row = 6
    ws.cell(row=row, column=1, value="By Severity").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Severity").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    for sev in ["BLOCKING", "WARNING", "INFO"]:
        row += 1
        count = sum(1 for f in findings if f["severity"] == sev)
        ws.cell(row=row, column=1, value=sev).fill = SEVERITY_FILL[sev]
        ws.cell(row=row, column=2, value=count)

    # By track
    row += 2
    ws.cell(row=row, column=1, value="By Track").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Track").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    track_names = {"A": "Spec Conformance", "B": "Code Quality & Security",
                   "C": "Doc Consistency", "D": "Test Suite", "E": "Forensic L4"}
    for t in ["A", "B", "C", "D", "E"]:
        row += 1
        count = sum(1 for f in findings if t in f["track"])
        ws.cell(row=row, column=1, value=f"Track {t} ({track_names[t]})")
        ws.cell(row=row, column=2, value=count)

    # By Round-1 status
    row += 2
    ws.cell(row=row, column=1, value="By Round-1 Status").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Status").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    for status, label in [("partial", "Partial (carry-over)"),
                           ("persistent", "Persistent (carry-over)"),
                           ("new", "New in Round 2")]:
        row += 1
        count = sum(1 for f in findings if f["round1_status"] == status)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)

    # By category
    row += 2
    ws.cell(row=row, column=1, value="By Category (top 10)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Category").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    cat_counts: dict[str, int] = {}
    for f in findings:
        cat_counts[f["category"]] = cat_counts.get(f["category"], 0) + 1
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1])[:10]:
        row += 1
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=count)

    # Total effort estimate
    row += 2
    ws.cell(row=row, column=1, value="Estimated Remediation Effort").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Severity").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Hours/Item").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="Total Hours").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    total_hours = 0
    for sev in ["BLOCKING", "WARNING", "INFO"]:
        row += 1
        count = sum(1 for f in findings if f["severity"] == sev)
        hours = count * EFFORT_HOURS[sev]
        total_hours += hours
        ws.cell(row=row, column=1, value=sev).fill = SEVERITY_FILL[sev]
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=EFFORT_HOURS[sev])
        ws.cell(row=row, column=4, value=hours)
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_hours).font = Font(bold=True)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    # === Sheet 2: Findings (full register) ===
    ws = wb.create_sheet("Findings")
    style_header(ws, FINDINGS_COLUMNS)
    for idx, f in enumerate(findings, 2):
        write_finding_row(ws, idx, f, FINDINGS_COLUMNS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FINDINGS_COLUMNS))}{len(findings) + 1}"
    ws.freeze_panes = "A2"

    # === Sheets 3-7: Per-track subsets ===
    for track_letter in ["A", "B", "C", "D", "E"]:
        ws = wb.create_sheet(f"Track {track_letter}")
        track_findings = [f for f in findings if track_letter in f["track"]]
        style_header(ws, FINDINGS_COLUMNS)
        for idx, f in enumerate(track_findings, 2):
            write_finding_row(ws, idx, f, FINDINGS_COLUMNS)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(FINDINGS_COLUMNS))}{len(track_findings) + 1}"
        ws.freeze_panes = "A2"

    # === Sheet 8: Round-1 Status ===
    ws = wb.create_sheet("Round1 Status")
    status_columns = [
        ("id", "ID", 10),
        ("severity", "Severity", 10),
        ("title", "Title", 60),
        ("round1_status", "Round 1 Status", 18),
        ("source_findings", "Source Findings (Round 2 tracks)", 50),
    ]
    style_header(ws, status_columns)
    status_order = {"partial": 0, "persistent": 1, "new": 2}
    sorted_findings = sorted(findings, key=lambda f: status_order.get(f["round1_status"], 3))
    for idx, f in enumerate(sorted_findings, 2):
        write_finding_row(ws, idx, f, status_columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(status_columns))}{len(findings) + 1}"
    ws.freeze_panes = "A2"

    # === Sheet 9: Remediation Backlog (priority-sorted) ===
    ws = wb.create_sheet("Remediation Backlog")
    backlog_columns = [
        ("id", "ID", 10),
        ("severity", "Severity", 10),
        ("category", "Category", 35),
        ("title", "Title", 50),
        ("suggested_fix", "Suggested Fix", 70),
        ("round1_status", "Carry-over/New", 14),
        ("effort_hours", "Effort (hrs)", 12),
    ]
    style_header(ws, backlog_columns)
    severity_order = {"BLOCKING": 0, "WARNING": 1, "INFO": 2}
    backlog = sorted(findings, key=lambda f: severity_order[f["severity"]])
    for idx, f in enumerate(backlog, 2):
        f_copy = dict(f)
        f_copy["effort_hours"] = EFFORT_HOURS[f["severity"]]
        write_finding_row(ws, idx, f_copy, backlog_columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(backlog_columns))}{len(backlog) + 1}"
    ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
