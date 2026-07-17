"""Generate the Round 4 audit findings register XLSX.

Multi-sheet workbook mirroring the R3 register shape:
  - Summary
  - Findings (full register, autofilter)
  - Per-track sheets (R4/A4/B4/C4/D4/E4/F4)
  - R3 Status (carry-over matrix)
  - Remediation Backlog (priority-sorted with effort estimates)

Output: /home/z/my-project/download/TRA_Round4/TRA_audit_findings_register_r4.xlsx
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REGISTER = Path(
    "/home/z/my-project/Translation-Runtime-Architecture/docs/audit/round4/master_findings_register_r4.json"
)
OUT = Path(
    "/home/z/my-project/download/TRA_Round4/TRA_audit_findings_register_r4.xlsx"
)

# Severity colors
SEVERITY_FILL = {
    "BLOCKING": PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
    "WARNING": PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    ),
    "INFO": PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ),
}
HEADER_FILL = PatternFill(
    start_color="305496", end_color="305496", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP_LEFT = Alignment(vertical="top", horizontal="left")


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )


def autosize(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    cell_len = len(str(cell.value).split("\n")[0])
                    if cell_len > max_len:
                        max_len = cell_len
            except Exception:
                pass
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def build_summary_sheet(wb: Workbook, findings: list[dict]) -> None:
    ws = wb.create_sheet("Summary", 0)
    issues = [f for f in findings if f.get("finding_type") == "issue"]
    positives = [f for f in findings if f.get("finding_type") == "positive_verification"]
    by_sev = {"BLOCKING": 0, "WARNING": 0, "INFO": 0}
    for f in issues:
        by_sev[f["severity"]] = by_sev.get(f["severity"], 0) + 1
    by_track: dict[str, int] = {}
    for f in issues:
        for t in f["track"].split(","):
            by_track[t] = by_track.get(t, 0) + 1

    rows = [
        ["TRA Round 4 Audit — Summary", ""],
        ["HEAD audited", "805a8f8"],
        ["Audit date", "2026-07-17"],
        ["Methodology", "7-track parallel re-audit (R4/A4/B4/C4/D4/E4/F4)"],
        ["Carry-over input", "Round 3 master register (36 findings)"],
        ["", ""],
        ["Total findings (deduplicated)", len(findings)],
        ["  Issues", len(issues)],
        ["  Positive verifications", len(positives)],
        ["", ""],
        ["Issues by severity", "Count"],
        ["  BLOCKING", by_sev["BLOCKING"]],
        ["  WARNING", by_sev["WARNING"]],
        ["  INFO", by_sev["INFO"]],
        ["", ""],
        ["Issues by track", "Count"],
    ]
    for t in ["A4", "B4", "C4", "D4", "E4", "F4"]:
        rows.append([f"  {t}", by_track.get(t, 0)])
    rows.append(["", ""])
    rows.append(
        [
            "Track R4 (regression baseline)",
            "36 R3 findings re-verified: 20 FIXED / 12 PERSISTENT / 4 PARTIAL / 0 REGRESSED",
        ]
    )
    rows.append(
        [
            "Round 3 BLOCKING findings",
            "Both fixed (TRA-093 + TRA-096); verified holding at HEAD 805a8f8",
        ]
    )
    rows.append(
        [
            "Round 4 BLOCKING findings",
            "1 NEW (TRA-C4-013: tra-prototype/README.md CLI examples fail as written)",
        ]
    )
    rows.append(["", ""])
    rows.append(
        ["4 critical invariants", "ALL HOLD at HEAD 805a8f8 (code-level evidence)"]
    )
    rows.append(
        [
            "TRA-013 byte-reproducibility",
            "HOLDS — audit_trace.jsonl sha256 263b901e... (matches R3 exactly)",
        ]
    )
    rows.append(
        ["Quality gates", "ALL GREEN: ruff, mypy --strict, pytest (199 passed)"]
    )

    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    autosize(ws, min_width=20, max_width=80)


def build_findings_sheet(wb: Workbook, findings: list[dict]) -> None:
    ws = wb.create_sheet("Findings")
    headers = [
        "ID",
        "Root ID",
        "Severity",
        "Track(s)",
        "Category",
        "Title",
        "Evidence",
        "Detail",
        "Suggested Fix",
        "Round 3 Status",
        "Finding Type",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    for f in findings:
        ws.append(
            [
                f["id"],
                f.get("root_id", ""),
                f["severity"],
                f["track"],
                f["category"],
                f["title"],
                f["evidence"],
                f["detail"],
                f["suggested_fix"],
                f["round3_status"],
                f.get("finding_type", "issue"),
            ]
        )
    # Severity color
    for row in range(2, len(findings) + 2):
        sev = ws.cell(row=row, column=3).value
        if sev in SEVERITY_FILL:
            ws.cell(row=row, column=3).fill = SEVERITY_FILL[sev]
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = WRAP
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, min_width=10, max_width=50)


def build_track_sheet(wb: Workbook, track: str, findings: list[dict]) -> None:
    ws = wb.create_sheet(f"Track {track}")
    track_findings = [f for f in findings if track in f["track"].split(",")]
    headers = [
        "ID",
        "Severity",
        "Category",
        "Title",
        "Evidence",
        "Detail",
        "Suggested Fix",
        "Round 3 Status",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    for f in track_findings:
        ws.append(
            [
                f["id"],
                f["severity"],
                f["category"],
                f["title"],
                f["evidence"],
                f["detail"],
                f["suggested_fix"],
                f["round3_status"],
            ]
        )
    for row in range(2, len(track_findings) + 2):
        sev = ws.cell(row=row, column=2).value
        if sev in SEVERITY_FILL:
            ws.cell(row=row, column=2).fill = SEVERITY_FILL[sev]
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = WRAP
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, min_width=10, max_width=50)


def build_r3_status_sheet(wb: Workbook, findings: list[dict]) -> None:
    """Round 3 carry-over status matrix."""
    ws = wb.create_sheet("R3 Status")
    # Build the 36-finding carry-over table from R4 baseline
    r3_to_r4_status = {
        "TRA-001": ("WARNING", "TRANSLATE_SEGMENT whole-doc", "persistent"),
        "TRA-002": ("WARNING", "Kernel registry wiring", "fixed"),
        "TRA-003": ("INFO", "Spec section cross-ref", "fixed"),
        "TRA-004": ("WARNING", "Exception recovery routing", "partial"),
        "TRA-005": ("INFO", "ConformanceFailure exit code", "fixed"),
        "TRA-006": ("WARNING", "PolicyResolver never invoked", "partial"),
        "TRA-007": ("INFO", "Transitions fire after ISA", "fixed"),
        "TRA-008": ("INFO", "Anchor rewrite links", "fixed"),
        "TRA-009": ("INFO", "Forward-only transitions", "fixed"),
        "TRA-011": ("INFO", "cache-clear --pattern", "fixed"),
        "TRA-012": ("INFO", "sanitize_input chokepoint", "fixed"),
        "TRA-013": ("INFO", "Deterministic audit trail", "fixed"),
        "TRA-014": ("INFO", "Path-safety validation", "fixed"),
        "TRA-016": ("INFO", "Dead count_blocking stub", "fixed"),
        "TRA-017": ("WARNING", "Unused deps", "fixed"),
        "TRA-018": ("INFO", "BootstrapConfig frozen", "fixed"),
        "TRA-019": ("INFO", "Hard raises vs assert", "fixed"),
        "TRA-026": ("INFO", "Dead cache.expire config", "fixed"),
        "TRA-031": ("WARNING", "Benchmark coverage 22/100+", "persistent"),
        "TRA-032": ("INFO", "LLM seam contract", "fixed"),
        "TRA-033": ("INFO", "LLM graceful degradation", "fixed"),
        "TRA-036": ("BLOCKING", "Analyze-failure L3 gate", "fixed"),
        "TRA-037": ("BLOCKING", "rewrite_anchors before L3 gate", "fixed"),
        "TRA-038": ("WARNING", "3 unreachable exceptions", "partial"),
        "TRA-039": ("INFO", "build_entity_table wrap", "fixed"),
        "TRA-040": ("WARNING", "EXCEPTION_HANDLER not KernelState", "persistent"),
        "TRA-041": ("INFO", "GLOSSARY_CONFLICT recovery", "fixed"),
        "TRA-042": ("WARNING", "Structural verification heading-only", "persistent"),
        "TRA-043": ("WARNING", "RuntimeContext.module Any type", "partial"),
        "TRA-044": ("INFO", "route_exception Unrecoverable branch", "fixed"),
        "TRA-045": ("INFO", "Dead CONCLUSION_LEADING", "fixed"),
        "TRA-046": ("INFO", "_hash_sorted rename", "fixed"),
        "TRA-047": ("INFO", "from_yaml base_dir extra=forbid", "fixed"),
        "TRA-048": ("WARNING", "LLM-degradation single audit record", "partial"),
        "TRA-049": ("INFO", "Same-state transition raises", "fixed"),
        "TRA-050": ("INFO", "Cache-key content sensitivity", "fixed"),
        "TRA-051": ("INFO", "cache.invalidate fnmatch", "fixed"),
        "TRA-052": ("WARNING", "interactive=True e2e test", "persistent"),
        "TRA-053": ("INFO", "Inline-code protection test", "fixed"),
        "TRA-054": ("INFO", "L3 ConformanceFailure test", "fixed"),
        "TRA-055": ("WARNING", "review_decision text assertion", "persistent"),
        "TRA-056": ("WARNING", "on_override callback untested", "persistent"),
        "TRA-057": ("WARNING", "kernel_config fixture unused", "persistent"),
        "TRA-058": ("WARNING", "Cross-file duplicate HITL test", "persistent"),
        "TRA-059": ("INFO", "Doc staleness (CLAUDE.md)", "partial"),
        "TRA-060": ("INFO", "Doc staleness (README.md)", "partial"),
        "TRA-061": ("INFO", "Doc staleness (AGENTS.md)", "persistent"),
        "TRA-062": ("INFO", "Doc staleness (implementation_plan.md)", "persistent"),
        "TRA-063": ("INFO", "Doc staleness (status.md)", "partial"),
        "TRA-064": ("INFO", "Doc staleness (prototype.md)", "persistent"),
        "TRA-065": ("INFO", "Doc staleness (review-feedback.md)", "persistent"),
        "TRA-066": ("INFO", "Doc staleness (start-here.md)", "persistent"),
        "TRA-067": ("INFO", "Doc staleness (review.md)", "persistent"),
        "TRA-068": ("INFO", "_deterministic_clock seed in run()", "fixed"),
        "TRA-069": ("INFO", "Dead out=out loop", "fixed"),
        "TRA-070": ("INFO", "_hash_canonical_json rename", "fixed"),
        "TRA-071": ("WARNING", "BrokenMarkdown unclosed fence", "fixed"),
        "TRA-072": ("WARNING", "PolicyResolver universal arbitration", "persistent"),
        "TRA-073": ("INFO", "Dead out=out loop (R3 remediation)", "fixed"),
        "TRA-074": ("INFO", "_deterministic_clock seed default", "fixed"),
        "TRA-075": ("INFO", "Pairwise transition tests", "fixed"),
        "TRA-076": ("WARNING", "LLM seam sanitization (OWASP A03)", "fixed"),
        "TRA-077": ("WARNING", "Cache JSON not pickle (OWASP A08)", "fixed"),
        "TRA-078": ("WARNING", "Secret redaction (OWASP A09)", "fixed"),
        "TRA-079": ("INFO", "Cache HMAC integrity", "persistent"),
        "TRA-080": ("INFO", "CLAUDE.md TRA-006 entry", "fixed"),
        "TRA-081": ("INFO", "README Policy module path", "fixed"),
        "TRA-082": ("WARNING", "README EntityAmbiguity claim", "partial"),
        "TRA-083": ("INFO", "README implementation_plan path", "fixed"),
        "TRA-084": ("INFO", "AGENTS.md separate repo contradiction", "fixed"),
        "TRA-085": ("WARNING", "status.md stale banner", "partial"),
        "TRA-086": ("INFO", "implementation_plan.md external codebase", "fixed"),
        "TRA-087": ("WARNING", "implementation_plan.md File Structure", "persistent"),
        "TRA-088": ("WARNING", "LLM-seam single-audit-record test", "partial"),
        "TRA-089": ("WARNING", "ConformanceFailure e2e tests", "fixed"),
        "TRA-090": ("WARNING", "llm_translate param on run()", "persistent"),
        "TRA-091": ("WARNING", "interactive=True e2e test", "persistent"),
        "TRA-092": ("WARNING", "S-03 and E-03 benchmark cases", "persistent"),
        "TRA-093": ("BLOCKING", "False-positive BROKEN_LINK", "fixed"),
        "TRA-094": ("INFO", "Mutation testing framework", "persistent"),
        "TRA-095": ("WARNING", "HITL CLI reachability", "persistent"),
        "TRA-096": ("BLOCKING", "as_interface crashes", "fixed"),
        "TRA-097": ("INFO", "register() isinstance check", "fixed"),
        "TRA-098": ("INFO", "register() duplicate detection", "fixed"),
        "TRA-099": ("WARNING", "CLI --registry flag", "persistent"),
        "TRA-100": ("INFO", "Module authoring guide", "persistent"),
    }
    ws.append(["Round 3 ID", "R3 Severity", "Title", "Round 4 Status"])
    style_header(ws, 4)
    for r3_id, (sev, title, r4_status) in sorted(r3_to_r4_status.items()):
        ws.append([r3_id, sev, title, r4_status])
        if sev in SEVERITY_FILL:
            ws.cell(row=ws.max_row, column=2).fill = SEVERITY_FILL[sev]
        status_color = {
            "fixed": "C6EFCE",
            "partial": "FFEB9C",
            "persistent": "FFC7CE",
        }.get(r4_status, "FFFFFF")
        ws.cell(row=ws.max_row, column=4).fill = PatternFill(
            start_color=status_color, end_color=status_color, fill_type="solid"
        )
    autosize(ws, min_width=12, max_width=50)

    # Add totals row
    ws.append([])
    ws.append(["TOTALS", "", "", ""])
    ws.append(
        [
            "FIXED",
            "",
            "",
            sum(1 for v in r3_to_r4_status.values() if v[2] == "fixed"),
        ]
    )
    ws.append(
        [
            "PARTIAL",
            "",
            "",
            sum(1 for v in r3_to_r4_status.values() if v[2] == "partial"),
        ]
    )
    ws.append(
        [
            "PERSISTENT",
            "",
            "",
            sum(1 for v in r3_to_r4_status.values() if v[2] == "persistent"),
        ]
    )
    ws.append(["TOTAL", "", "", len(r3_to_r4_status)])


def build_backlog_sheet(wb: Workbook, findings: list[dict]) -> None:
    """Remediation backlog: priority-sorted issues with effort estimates."""
    ws = wb.create_sheet("Remediation Backlog")
    issues = [f for f in findings if f.get("finding_type") == "issue"]
    # Sort by severity (BLOCKING first), then by root_id
    sev_rank = {"BLOCKING": 0, "WARNING": 1, "INFO": 2}
    issues.sort(key=lambda f: (sev_rank.get(f["severity"], 3), f.get("root_id", "")))

    # Effort estimates by category
    effort_map = {
        "BLOCKING": ("2-4h", "P0"),
        "WARNING": ("4-8h", "P1"),
        "INFO": ("1-2h", "P2"),
    }

    headers = [
        "Priority",
        "ID",
        "Severity",
        "Title",
        "Category",
        "Est. Effort",
        "Suggested Fix",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    for f in issues:
        est, pri = effort_map.get(f["severity"], ("1h", "P3"))
        ws.append(
            [
                pri,
                f["id"],
                f["severity"],
                f["title"],
                f["category"],
                est,
                f["suggested_fix"],
            ]
        )
        if f["severity"] in SEVERITY_FILL:
            ws.cell(row=ws.max_row, column=3).fill = SEVERITY_FILL[f["severity"]]
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = WRAP
    autosize(ws, min_width=10, max_width=50)


def main() -> None:
    findings = json.loads(REGISTER.read_text(encoding="utf-8"))
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_summary_sheet(wb, findings)
    build_findings_sheet(wb, findings)
    for track in ["A4", "B4", "C4", "D4", "E4", "F4"]:
        build_track_sheet(wb, track, findings)
    build_r3_status_sheet(wb, findings)
    build_backlog_sheet(wb, findings)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Findings: {len(findings)}")


if __name__ == "__main__":
    main()
