"""Generate multi-sheet XLSX findings register for Round 3 audit."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REGISTER = Path("/home/z/my-project/download/TRA_Round3/master_findings_register_r3.json")
OUT = Path("/home/z/my-project/download/TRA_Round3/TRA_audit_findings_register_r3.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BLOCKING_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
INFO_FILL = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("ID", 12),
    ("Severity", 12),
    ("Category", 35),
    ("Track", 10),
    ("Title", 50),
    ("Evidence", 60),
    ("Detail", 70),
    ("Suggested Fix", 60),
    ("Round 2 Status", 15),
    ("Source Findings", 35),
]


def style_header(ws) -> None:
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def severity_fill(severity: str) -> PatternFill | None:
    if severity == "BLOCKING":
        return BLOCKING_FILL
    if severity == "WARNING":
        return WARNING_FILL
    if severity == "INFO":
        return INFO_FILL
    return None


def write_findings_sheet(ws, findings: list[dict]) -> None:
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    style_header(ws)
    for f in findings:
        row = [
            f["id"],
            f["severity"],
            f["category"],
            f["track"],
            f["title"],
            f["evidence"],
            f["detail"],
            f["suggested_fix"],
            f["round2_status"],
            ", ".join(f.get("source_findings", [])),
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill = severity_fill(f["severity"])
        if fill:
            ws.cell(row=row_idx, column=2).fill = fill
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP


def write_summary_sheet(ws, findings: list[dict]) -> None:
    ws.append(["TRA Round 3 Audit — Summary"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total findings", len(findings)])
    ws.append([])
    # By severity
    ws.append(["By Severity"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    sev = {}
    for f in findings:
        sev[f["severity"]] = sev.get(f["severity"], 0) + 1
    for s in ["BLOCKING", "WARNING", "INFO"]:
        ws.append([s, sev.get(s, 0)])
    ws.append([])
    # By track
    ws.append(["By Track"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    track = {}
    for f in findings:
        t = f["track"].split(",")[0]
        track[t] = track.get(t, 0) + 1
    for t in ["R3", "A3", "B3", "C3", "D3", "E3", "F3"]:
        ws.append([t, track.get(t, 0)])
    ws.append([])
    # By Round 2 status
    ws.append(["By Round 2 Status"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    r2 = {}
    for f in findings:
        r2[f["round2_status"]] = r2.get(f["round2_status"], 0) + 1
    for s in ["fixed", "persistent", "partial", "new", "n/a"]:
        ws.append([s, r2.get(s, 0)])
    ws.append([])
    # BLOCKING findings list
    ws.append(["BLOCKING Findings (immediate attention)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="FF0000")
    for f in findings:
        if f["severity"] == "BLOCKING":
            ws.append([f["id"], f["title"], f["track"]])
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15


def write_track_sheet(ws, findings: list[dict], track: str) -> None:
    ws.append([f"Track {track} Findings"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    track_findings = [f for f in findings if f["track"].split(",")[0] == track]
    write_findings_sheet(ws, track_findings)


def write_remediation_backlog(ws, findings: list[dict]) -> None:
    ws.append(["Remediation Backlog — Priority Sorted"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    headers = ["ID", "Severity", "Title", "Suggested Fix", "Est. Effort (hrs)"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(ws.max_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    # Sort: BLOCKING first, then WARNING, then INFO
    sev_order = {"BLOCKING": 0, "WARNING": 1, "INFO": 2}
    sorted_f = sorted(findings, key=lambda f: (sev_order.get(f["severity"], 9), f["id"]))
    effort = {"BLOCKING": 4, "WARNING": 2, "INFO": 1}
    for f in sorted_f:
        ws.append([
            f["id"],
            f["severity"],
            f["title"],
            f["suggested_fix"][:200] + "..." if len(f["suggested_fix"]) > 200 else f["suggested_fix"],
            effort.get(f["severity"], 1),
        ])
        row_idx = ws.max_row
        fill = severity_fill(f["severity"])
        if fill:
            ws.cell(row=row_idx, column=2).fill = fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 15
    ws.freeze_panes = "A3"
    # Total effort
    ws.append([])
    total = sum(effort.get(f["severity"], 1) for f in sorted_f)
    ws.append(["Total estimated effort", f"{total} hours ({total/8:.1f} person-days)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)


def main() -> None:
    findings = json.loads(REGISTER.read_text(encoding="utf-8"))
    wb = Workbook()
    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    write_summary_sheet(ws_sum, findings)
    # All findings sheet
    ws_all = wb.create_sheet("Findings")
    write_findings_sheet(ws_all, findings)
    # Per-track sheets
    for track in ["R3", "A3", "B3", "C3", "D3", "E3", "F3"]:
        ws_t = wb.create_sheet(f"Track {track}")
        write_track_sheet(ws_t, findings, track)
    # Round 2 status sheet
    ws_r2 = wb.create_sheet("Round2 Status")
    ws_r2.append(["Round 2 Finding Status at HEAD b783745"])
    ws_r2.cell(1, 1).font = Font(bold=True, size=14)
    ws_r2.append([])
    r2_status = {}
    for f in findings:
        status = f["round2_status"]
        if status not in r2_status:
            r2_status[status] = []
        r2_status[status].append(f["id"])
    for status in ["fixed", "persistent", "partial", "new", "n/a"]:
        ids = r2_status.get(status, [])
        ws_r2.append([status, len(ids), ", ".join(ids)])
    ws_r2.column_dimensions["A"].width = 20
    ws_r2.column_dimensions["B"].width = 10
    ws_r2.column_dimensions["C"].width = 80
    # Remediation backlog
    ws_back = wb.create_sheet("Remediation Backlog")
    write_remediation_backlog(ws_back, findings)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
